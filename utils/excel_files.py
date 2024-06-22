from openpyxl import load_workbook, Workbook

def read_excel_file(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    datas = {}
    for row in ws.iter_rows(values_only= True):
        datas[row[0]] = {'definition': row[1], 'learned': False}
    return datas

def write_excel_file(file_name, datas):
    wb = Workbook()
    ws = wb.active
    for k , v in datas.items():
        ws.append([k, v['definition']])
    wb.save(file_name)
    

if __name__ == '__main__':
    datas = read_excel_file('files\\words.xlsx')
    print(datas)
    write_excel_file('test.xlsx', datas)
